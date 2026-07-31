"""Export MuleSoft-vs-AI review data to Excel for analyst spot checks.

One row per household member, grouped under a merged Case ID cell:

    Case ID | MuleSoft: Full name, Address, Phone, Email, SSN, Income |
            | AI Audit: Full name, Address, Phone, Email, SSN, Income

SSNs are UNMASKED in this workbook (that is its purpose — the audit UI and
Salesforce findings stay masked). Both sides can only show what their system
holds: Salesforce stores many SSNs masked at rest, and AI extractions made
before the capture-as-printed change (2026-07-31) are permanently masked.

The address comes from the Case record (property address + unit) and is the
same physical unit for both sides; the AI side repeats it since extraction
does not capture street addresses.

Usage:
    .venv/bin/python export_review_data.py [--out FILE] [--limit N]

The output defaults to /var/data/audit_review_export_<date>.xlsx. Handle
the file as sensitive: it contains unmasked SSNs.
"""

import argparse
import datetime as dt
import json
import sqlite3
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.core.config import Settings
from app.services.salesforce.client import SalesforceClient

DB_PATH = "/var/data/audit_jobs.db"

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

SIDE_COLS = ["Full name", "DOB", "Address", "Phone", "Email", "SSN", "Income"]
# Column layout: A = Case ID, B-H = MuleSoft, I = spacer, J-P = AI Audit.
_MS_BASE, _SPACER, _AI_BASE = 2, 9, 10


def _full_name(first, last):
    return " ".join(p for p in (first, last) if p) or None


def _fmt_money(v):
    try:
        return f"${float(v):,.2f}"
    except (TypeError, ValueError):
        return v or None


def _case_addresses(sf, case_ids):
    """Batch-fetch property address + unit per case from Salesforce."""
    out = {}
    ids = list(case_ids)
    for i in range(0, len(ids), 200):
        chunk = ", ".join(f"'{cid}'" for cid in ids[i:i + 200])
        q = (
            "SELECT Id, Property_Address__c, Property_City__c, Property_State__c, "
            f"Property_Zip__c, Unit_Number__c FROM Case WHERE Id IN ({chunk})"
        )
        for r in sf.sf.query_all(q)["records"]:
            parts = [r.get("Property_Address__c")]
            city = r.get("Property_City__c")
            state = r.get("Property_State__c")
            zipc = r.get("Property_Zip__c")
            if city or state or zipc:
                parts.append(", ".join(p for p in (city, f"{state or ''} {zipc or ''}".strip()) if p))
            unit = r.get("Unit_Number__c")
            if unit:
                parts.append(f"Unit {unit}")
            out[r["Id"]] = ", ".join(p for p in parts if p) or None
    return out


def _mulesoft_members(snapshot):
    """(full_name, phone, email, ssn, income) per member from a snapshot."""
    rows = []
    if not snapshot:
        return rows
    members = snapshot.get("members") or []
    # Per-member income: sum of that member's worksheet gross amounts.
    income_by_member = defaultdict(float)
    for w in snapshot.get("income") or []:
        member_ref = w.get("Household_Member__c")
        amt = w.get("Gross_Member_Income__c") or w.get("Self_Declared_Annual_Amount__c")
        try:
            income_by_member[member_ref] += float(amt or 0)
        except (TypeError, ValueError):
            pass
    for m in members:
        income = income_by_member.get(m.get("Id"))
        rows.append({
            "name": m.get("Full_Name__c") or _full_name(m.get("First_Name__c"), m.get("Last_Name__c")),
            "dob": m.get("DOB__c"),
            "phone": m.get("Tenant_Phone__c"),
            "email": m.get("Tenant_Email__c"),
            "ssn": m.get("SSN__c"),
            "income": _fmt_money(income) if income else None,
        })
    return rows


def _ai_members(extraction):
    """(full_name, phone, email, ssn, income) per member from extraction."""
    rows = []
    if not extraction:
        return rows
    hh = (extraction.get("household_demographics") or {}).get("houseHold") or []
    # Per-member income: sum of primary income calculations by given name.
    calcs = extraction.get("income_calculations") or []
    income_by_name = defaultdict(float)
    for c in calcs:
        details = c.get("details") or ""
        if "[audit]" in details or "[historical]" in details:
            continue
        name = (c.get("memberName") or "").strip().lower()
        try:
            income_by_name[name] += float(c.get("annualIncome") or 0)
        except (TypeError, ValueError):
            pass
    for m in hh:
        name = _full_name(m.get("FirstName"), m.get("LastName"))
        income = None
        if name:
            key = name.strip().lower()
            # tolerate name variants: exact, else first-token + shared token
            if key in income_by_name:
                income = income_by_name[key]
            else:
                first = key.split()[0]
                for cand, amt in income_by_name.items():
                    if cand.split() and cand.split()[0] == first:
                        income = amt
                        break
        rows.append({
            "name": name,
            "dob": m.get("DOB"),
            "phone": m.get("phone"),
            "email": m.get("email"),
            "ssn": m.get("socialSecurityNumber"),
            "income": _fmt_money(income) if income else None,
        })
    return rows


def _ssn4(v):
    digits = "".join(ch for ch in (v or "") if ch.isdigit())
    return digits[-4:] if len(digits) >= 4 else None


def _align(ms_rows, ai_rows):
    """Reorder AI rows so the same person sits beside their MuleSoft row.

    Matches by SSN last-4 first, then given name; unmatched AI members
    keep their order after the matched ones."""
    remaining = list(ai_rows)
    aligned = []
    for ms in ms_rows:
        hit = None
        for ai in remaining:
            if _ssn4(ms.get("ssn")) and _ssn4(ms.get("ssn")) == _ssn4(ai.get("ssn")):
                hit = ai
                break
        if hit is None:
            ms_first = (ms.get("name") or "").lower().split()
            for ai in remaining:
                ai_first = (ai.get("name") or "").lower().split()
                if ms_first and ai_first and ms_first[0] == ai_first[0]:
                    hit = ai
                    break
        if hit is not None:
            remaining.remove(hit)
        aligned.append(hit or {})
    aligned.extend(remaining)
    # drop trailing all-empty placeholders beyond AI's real member count
    while aligned and not aligned[-1]:
        aligned.pop()
    return aligned


def build_workbook(out_path, limit=None):
    con = sqlite3.connect(DB_PATH)
    q = (
        "SELECT case_id, case_number, extraction_result, mulesoft_snapshot "
        "FROM audit_jobs WHERE state='done' AND extraction_result IS NOT NULL "
        "ORDER BY completed_at DESC"
    )
    if limit:
        q += f" LIMIT {int(limit)}"
    jobs = con.execute(q).fetchall()
    print(f"{len(jobs)} completed cases to export")

    sf = SalesforceClient(Settings())
    addresses = _case_addresses(sf, [j[0] for j in jobs])
    print(f"{len(addresses)} case addresses fetched")

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    # Header rows per the requested style: merged side bands over sub-headers.
    n_cols = len(SIDE_COLS)
    ws.cell(row=1, column=1)
    ws.merge_cells(start_row=1, start_column=_MS_BASE, end_row=1,
                   end_column=_MS_BASE + n_cols - 1)
    ws.cell(row=1, column=_MS_BASE, value="MuleSoft").font = BOLD
    ws.cell(row=1, column=_MS_BASE).alignment = CENTER
    ws.merge_cells(start_row=1, start_column=_AI_BASE, end_row=1,
                   end_column=_AI_BASE + n_cols - 1)
    ws.cell(row=1, column=_AI_BASE, value="AI Audit").font = BOLD
    ws.cell(row=1, column=_AI_BASE).alignment = CENTER

    ws.cell(row=2, column=1, value="Case ID").font = BOLD
    for i, name in enumerate(SIDE_COLS):
        for base in (_MS_BASE, _AI_BASE):
            c = ws.cell(row=2, column=base + i, value=name)
            c.font = BOLD
            c.alignment = CENTER
            c.border = BORDER

    row = 3
    for case_id, case_number, ex_json, snap_json in jobs:
        try:
            extraction = json.loads(ex_json)
        except (TypeError, json.JSONDecodeError):
            extraction = None
        try:
            snapshot = json.loads(snap_json) if snap_json else None
        except (TypeError, json.JSONDecodeError):
            snapshot = None

        ms_rows = _mulesoft_members(snapshot)
        ai_rows = _align(ms_rows, _ai_members(extraction))
        n = max(len(ms_rows), len(ai_rows), 1)
        address = addresses.get(case_id)

        start = row
        for i in range(n):
            ms = ms_rows[i] if i < len(ms_rows) else {}
            ai = ai_rows[i] if i < len(ai_rows) else {}
            vals = {1: case_number}
            for base, side in ((_MS_BASE, ms), (_AI_BASE, ai)):
                vals[base] = side.get("name")
                vals[base + 1] = side.get("dob")
                vals[base + 2] = address if side else None
                vals[base + 3] = side.get("phone")
                vals[base + 4] = side.get("email")
                vals[base + 5] = side.get("ssn")
                vals[base + 6] = side.get("income")
            for col, v in vals.items():
                c = ws.cell(row=row, column=col, value=v)
                if col != _SPACER:
                    c.border = BORDER
            row += 1
        if n > 1:
            ws.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
            ws.cell(row=start, column=1).alignment = Alignment(vertical="top")

    # Column widths
    widths = {1: 12, _SPACER: 2}
    for base in (_MS_BASE, _AI_BASE):
        for i, w in enumerate((22, 12, 38, 15, 28, 13, 14)):
            widths[base + i] = w
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"
    wb.save(out_path)
    print(f"wrote {out_path} ({row - 3} member rows, {len(jobs)} cases)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--out",
        default=f"/var/data/audit_review_export_{dt.date.today():%Y%m%d}.xlsx",
    )
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()
    build_workbook(args.out, args.limit)
